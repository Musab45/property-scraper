from __future__ import annotations

import os
import csv
import random
import re
import threading
import time
from urllib.parse import urlparse
from typing import Callable, Optional

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import undetected_chromedriver as uc
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.common.by import By

from contact_phone_extractor import reveal_and_extract_agent_phone
from login_manager import login_commercialguru, login_propertyguru

from scraper import (
    ScraperConfig,
    STANDARD_CSV_FIELDS,
    UC_DRIVER_CREATE_LOCK,
    extract_baths_from_soup,
    extract_beds_from_soup,
    extract_land_size_from_soup,
    detect_installed_chrome_major,
    extract_psf_from_soup,
)

DETAIL_FIELDS = STANDARD_CSV_FIELDS


# ── File reading helpers ──────────────────────────────────────────────────────

def read_urls_from_file(file_path: str, url_column: str) -> list[str]:
    """Read listing URLs from an Excel (.xlsx) or CSV (.csv) file.

    Looks for *url_column* (case-insensitive) and returns every cell that
    starts with ``http``.
    """
    ext = os.path.splitext(file_path)[1].lower()
    urls: list[str] = []

    if ext == ".xlsx":
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        first_row = next(rows_iter, None)
        if first_row is None:
            wb.close()
            return []
        headers = [str(h) if h is not None else "" for h in first_row]
        col_idx = _find_col_index(headers, url_column)
        if col_idx < 0:
            wb.close()
            raise ValueError(
                f"Column '{url_column}' not found in file.\n"
                f"Available columns: {headers}"
            )
        for row in rows_iter:
            if col_idx < len(row):
                val = row[col_idx]
                if val and str(val).strip().lower().startswith("http"):
                    urls.append(str(val).strip())
        wb.close()

    elif ext == ".csv":
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            field_names: list[str] = list(reader.fieldnames or [])
            actual_col = _find_col_name(field_names, url_column)
            if actual_col is None:
                raise ValueError(
                    f"Column '{url_column}' not found in CSV.\n"
                    f"Available columns: {field_names}"
                )
            for row in reader:
                val = row.get(actual_col, "").strip()
                if val.lower().startswith("http"):
                    urls.append(val)
    else:
        raise ValueError(
            f"Unsupported file type '{ext}'. Please use .xlsx or .csv."
        )

    return urls


def _find_col_index(headers: list[str], target: str) -> int:
    t = target.strip().lower()
    for i, h in enumerate(headers):
        if h.strip().lower() == t:
            return i
    return -1


def _find_col_name(names: list[str], target: str) -> Optional[str]:
    t = target.strip().lower()
    for n in names:
        if n.strip().lower() == t:
            return n
    return None


# ── Scraper ───────────────────────────────────────────────────────────────────

class DirectListingScraper:
    """Scrapes individual listing detail pages from a given list of direct URLs.

    Each URL is opened in Chrome, the page HTML is parsed, and the extracted
    data is written as a single row in an output Excel file.  One Excel file is
    produced per run (all listings merged into one sheet).
    """

    # Classes that indicate a listing detail page has loaded
    _WAIT_CLASSES = [
        "amount",
        "price-amount",
        "listing-specs",
        "listing-amenities",
        "amenities",
        "property-address",
        "agent-info",
        "listed-by",
    ]

    def __init__(
        self,
        urls: list[str],
        config: ScraperConfig,
        log_callback: Optional[Callable[[str], None]] = None,
        progress_callback: Optional[Callable[[dict], None]] = None,
        stop_requested: Optional[Callable[[], bool]] = None,
    ):
        self.urls = urls
        self.config = config
        self.log = log_callback or (lambda _: None)
        self.progress = progress_callback or (lambda _: None)
        self.stop_requested = stop_requested or (lambda: False)
        self.error_count = 0
        self.start_time = 0.0
        self._driver: uc.Chrome | None = None
        self._force_stop = False
        self._driver_lock = threading.Lock()

    # ── Stop handling ─────────────────────────────────────────────────────────

    def request_stop(self, force: bool = False) -> None:
        self._force_stop = True
        if not force:
            return
        with self._driver_lock:
            driver = self._driver
        if driver is None:
            return
        try:
            if self._is_driver_alive(driver):
                driver.quit()
        except Exception:
            pass
        finally:
            with self._driver_lock:
                if self._driver is driver:
                    self._driver = None

    def _should_stop(self) -> bool:
        return self._force_stop or self.stop_requested()

    def _sleep_with_stop(self, seconds: float) -> None:
        deadline = time.time() + seconds
        while time.time() < deadline:
            if self._should_stop():
                raise RuntimeError("Stop requested. Aborting scrape.")
            time.sleep(0.1)

    def _polite_wait(self, min_s: float = 2.0, max_s: float = 4.0) -> None:
        self._sleep_with_stop(random.uniform(min_s, max_s))

    @staticmethod
    def _format_launch_error(exc: Exception) -> str:
        text = str(exc).strip()
        if "Stacktrace:" in text:
            text = text.split("Stacktrace:", 1)[0].strip()
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        if not lines:
            return exc.__class__.__name__
        return " | ".join(lines[:2])

    # ── Driver management ─────────────────────────────────────────────────────

    def _await_first_window(self, driver: uc.Chrome, timeout: float = 5.0) -> bool:
        deadline = time.time() + timeout
        recovery_attempted = False
        while time.time() < deadline:
            if self._should_stop():
                raise RuntimeError("Stop requested.")
            try:
                handles = driver.window_handles
                if handles:
                    try:
                        driver.switch_to.window(handles[0])
                    except Exception:
                        pass
                    return True
            except Exception:
                pass

            if not recovery_attempted:
                recovery_attempted = True
                try:
                    driver.switch_to.new_window("tab")
                    continue
                except Exception:
                    pass
            time.sleep(0.1)
        return False

    def _is_driver_alive(self, driver: uc.Chrome) -> bool:
        try:
            _ = driver.window_handles
            return True
        except Exception:
            return False

    def _create_driver(self) -> uc.Chrome:
        attempts = self.config.retries + 1
        last_error: Exception | None = None
        preferred_major = self.config.chrome_major or detect_installed_chrome_major()

        for attempt in range(1, attempts + 1):
            if self._should_stop():
                raise RuntimeError("Stop requested before browser init.")
            self.log(f"Starting browser ({attempt}/{attempts})...")
            driver: uc.Chrome | None = None
            try:
                options = uc.ChromeOptions()
                options.page_load_strategy = "eager"
                options.add_argument("--no-sandbox")
                options.add_argument("--disable-dev-shm-usage")
                options.add_argument("--disable-blink-features=AutomationControlled")
                if self.config.headless:
                    options.add_argument("--headless=new")

                with UC_DRIVER_CREATE_LOCK:
                    if preferred_major is not None:
                        self.log(f"Using Chrome major {preferred_major}.")
                        driver = uc.Chrome(
                            options=options,
                            use_subprocess=True,
                            version_main=preferred_major,
                        )
                    else:
                        driver = uc.Chrome(options=options, use_subprocess=True)

                if not self._await_first_window(driver):
                    raise RuntimeError("No active Chrome window became available.")
                driver.set_window_size(1366, 900)
                with self._driver_lock:
                    self._driver = driver
                return driver

            except Exception as exc:
                last_error = exc
                mismatch = re.search(r"Current browser version is\s+(\d+)\.", str(exc))
                if mismatch:
                    preferred_major = int(mismatch.group(1))
                    self.log(f"Re-detected Chrome major {preferred_major} from error; retrying.")
                if attempt < attempts:
                    self.log(f"Browser launch failed: {self._format_launch_error(exc)}. Retrying...")
                    try:
                        if driver:
                            driver.quit()
                    except Exception:
                        pass
                    self._sleep_with_stop(1.5)

        raise RuntimeError(
            f"Failed to init browser after {attempts} attempts: {last_error}"
        )

    def _wait_for_any_class(
        self, driver: uc.Chrome, class_names: list[str], timeout: int
    ) -> bool:
        deadline = time.time() + timeout
        while time.time() < deadline:
            if self._should_stop():
                raise RuntimeError("Stop requested.")
            if not self._is_driver_alive(driver):
                raise RuntimeError("Browser was closed.")
            for cls in class_names:
                try:
                    if driver.find_elements(By.CLASS_NAME, cls):
                        return True
                except Exception:
                    pass
            time.sleep(0.2)
        return False

    def _navigate_to_url(self, driver: uc.Chrome, url: str) -> bool:
        attempts = self.config.retries + 1
        for attempt in range(1, attempts + 1):
            if self._should_stop():
                raise RuntimeError("Stop requested.")
            if not self._is_driver_alive(driver):
                raise RuntimeError("Browser was closed.")
            try:
                driver.set_page_load_timeout(self.config.timeout_sec)
                driver.get(url)
            except TimeoutException:
                pass
            except WebDriverException:
                if self._should_stop():
                    raise RuntimeError("Stop requested.")
                raise RuntimeError("Browser closed during navigation.")
            except Exception:
                if attempt == attempts:
                    return False
                continue

            if self._wait_for_any_class(driver, self._WAIT_CLASSES, self.config.timeout_sec):
                self._polite_wait()
                return True

            if attempt < attempts:
                self.log(f"Retrying load ({attempt}/{attempts - 1}): {url}")

        return False

    # ── HTML parsing ──────────────────────────────────────────────────────────

    @staticmethod
    def _find_text(soup: BeautifulSoup, selectors: list[str]) -> Optional[str]:
        for sel in selectors:
            try:
                elem = soup.select_one(sel)
                if elem:
                    text = elem.get_text(separator=" ", strip=True)
                    if text:
                        return text
            except Exception:
                continue
        return None

    @staticmethod
    def _extract_meta_table(soup: BeautifulSoup) -> dict[str, str]:
        """Return a dict of label → value from the listing spec/meta table."""
        result: dict[str, str] = {}
        # Cover both PropertyGuru and CommercialGuru meta table structures
        for item in soup.select(
            ".listing-specifications__item, "
            ".meta-table__item, "
            ".listing-amenities__item"
        ):
            label_elem = item.select_one(
                ".listing-specifications__label, "
                ".meta-table__item__label, "
                ".meta-table__item__wrapper__label, "
                ".amenity-label"
            )
            value_elem = item.select_one(
                ".listing-specifications__value, "
                ".meta-table__item__value, "
                ".meta-table__item__wrapper__value, "
                ".amenity-value"
            )
            if label_elem and value_elem:
                k = label_elem.get_text(strip=True)
                v = value_elem.get_text(strip=True)
                if k and v:
                    result[k] = v
        return result

    @staticmethod
    def _extract_size_sqft(soup: BeautifulSoup) -> Optional[str]:
        for da_id in ("area-amenity", "floor-area-amenity", "land-area-amenity"):
            wrapper = soup.find("div", attrs={"da-id": da_id})
            if wrapper:
                for p in wrapper.find_all("p"):
                    text = p.get_text(strip=True)
                    if "sqft" in text.lower():
                        return text
        # Regex fallback anywhere on the page
        m = re.search(r"([\d,]+\s*sqft)", soup.get_text(separator=" "), re.IGNORECASE)
        return m.group(1).strip() if m else None

    @staticmethod
    def _extract_land_size_value(soup: BeautifulSoup) -> Optional[str]:
        texts: list[str] = []
        for da_id in ("area-amenity", "floor-area-amenity", "land-area-amenity"):
            wrapper = soup.find("div", attrs={"da-id": da_id})
            if not wrapper:
                continue
            for p in wrapper.find_all("p"):
                text = p.get_text(strip=True)
                if text:
                    texts.append(text)

        # Most reliable pattern: "1,706 sqft" or similar.
        for text in texts:
            match = re.search(r"([\d,]+)\s*(?:sq\s*ft|sqft)\b", text, re.IGNORECASE)
            if match:
                return match.group(1).strip()

        # Some pages split number and unit into separate nodes.
        for text in texts:
            if re.fullmatch(r"[\d,]+", text):
                return text

        # Final fallback from full page text.
        match = re.search(
            r"([\d,]+)\s*(?:sq\s*ft|sqft)\b",
            soup.get_text(separator=" "),
            re.IGNORECASE,
        )
        return match.group(1).strip() if match else None

    def _extract_tenure_value(self, soup: BeautifulSoup) -> Optional[str]:
        meta = self._extract_meta_table(soup)
        for key, value in meta.items():
            if "tenure" in key.lower() and value:
                return value

        # If explicit tenure is missing, keep parity with URL search behavior by
        # falling back to property type-like values.
        for key, value in meta.items():
            if key.lower() in {"property type", "type"} and value:
                return value

        elem = soup.select_one("[da-id='tenure-value'], .meta-table__item__wrapper__value")
        if elem:
            text = elem.get_text(strip=True)
            if text:
                return text
        return None

    @staticmethod
    def _extract_psf(soup: BeautifulSoup) -> Optional[str]:
        return extract_psf_from_soup(soup)

    @staticmethod
    def _extract_phone(soup: BeautifulSoup) -> Optional[str]:
        for a in soup.find_all("a", href=re.compile(r"^tel:")):
            num = a["href"].replace("tel:", "").strip()
            if num:
                return num
        for sel in (
            '[da-id="agent-phone"]',
            ".agent-phone",
            ".phone-number",
            ".contact-phone",
        ):
            elem = soup.select_one(sel)
            if elem:
                text = elem.get_text(strip=True)
                if text:
                    return text
        return None

    @staticmethod
    def _extract_email(soup: BeautifulSoup) -> Optional[str]:
        for a in soup.find_all("a", href=re.compile(r"^mailto:")):
            addr = a["href"].replace("mailto:", "").strip()
            if addr:
                return addr
        for sel in (
            '[da-id="agent-email"]',
            ".agent-email",
            ".contact-email",
        ):
            elem = soup.select_one(sel)
            if elem:
                text = elem.get_text(strip=True)
                if text:
                    return text
        return None

    def _parse_listing_page(self, html: str, url: str) -> dict:
        raise NotImplementedError("Use site-specific parsing methods.")

    def _parse_propertyguru_listing_page(self, html: str, url: str) -> dict:
        soup = BeautifulSoup(html, "html.parser")

        asking_price = None
        price_elem = soup.find("h2", class_="amount")
        if price_elem:
            asking_price = price_elem.get_text(strip=True)

        psf = extract_psf_from_soup(soup)
        beds = extract_beds_from_soup(soup)
        baths = extract_baths_from_soup(soup)
        land_size = extract_land_size_from_soup(soup)

        mrt_distance = None
        mrt_elem = soup.find("p", class_="mrt-distance__text")
        if mrt_elem:
            mrt_distance = mrt_elem.get_text(strip=True)

        tenure = self._extract_tenure_value(soup)

        agent_name = None
        agent_elem = soup.find("div", class_="agent-name")
        if agent_elem:
            agent_name = agent_elem.get_text(strip=True)

        district = None
        address_elem = soup.select_one('[da-id="property-address"], .listing-address')
        if address_elem:
            district = address_elem.get_text(strip=True) or None

        return {
            "URL": url,
            "District": district,
            "Asking Price": asking_price,
            "Beds": beds,
            "Baths": baths,
            "PSF": psf,
            "Nearest MRT + Distance": mrt_distance,
            "Land Size": land_size,
            "Tenure": tenure,
            "Agent Name": agent_name,
            "Agent Phone Number": self._extract_phone(soup),
        }

    def _parse_commercialguru_listing_page(self, html: str, url: str) -> dict:
        soup = BeautifulSoup(html, "html.parser")

        asking_price = None
        for selector in ['[da-id="price-amount"]', "h2.amount"]:
            elem = soup.select_one(selector)
            if elem:
                asking_price = elem.get_text(strip=True)
                break

        psf = extract_psf_from_soup(soup)
        beds = extract_beds_from_soup(soup)
        baths = extract_baths_from_soup(soup)
        land_size = extract_land_size_from_soup(soup)

        nearest_mrt_distance = None
        mrt_elem = soup.select_one('[da-id="mrt-distance-text"], p.mrt-distance__text')
        if mrt_elem:
            nearest_mrt_distance = mrt_elem.get_text(strip=True)

        tenure = self._extract_tenure_value(soup)

        agent_name = None
        for selector in [".agent-name", ".listed-by__agent-name"]:
            agent_elem = soup.select_one(selector)
            if agent_elem:
                agent_name = agent_elem.get_text(strip=True)
                break

        district = None
        address_elem = soup.select_one('[da-id="property-address"]')
        if address_elem and address_elem.get_text(strip=True):
            district = address_elem.get_text(strip=True)

        return {
            "URL": url,
            "District": district,
            "Asking Price": asking_price,
            "Beds": beds,
            "Baths": baths,
            "PSF": psf,
            "Nearest MRT + Distance": nearest_mrt_distance,
            "Land Size": land_size,
            "Tenure": tenure,
            "Agent Name": agent_name,
            "Agent Phone Number": self._extract_phone(soup),
        }

    # ── Main run loop ─────────────────────────────────────────────────────────

    def run(self) -> dict:
        self.start_time = time.time()
        total = len(self.urls)
        self.log(f"Direct listing scrape — {total} URL(s) to process")

        driver = self._create_driver()

        # Combined Excel — one row per listing
        wb = Workbook()
        ws = wb.active
        ws.title = "Listings"
        ws.append(DETAIL_FIELDS)

        processed = 0
        try:
            logged_in = {"propertyguru": False, "commercialguru": False}

            def _get_parser_for_url(listing_url: str):
                host = urlparse(listing_url).netloc.lower()
                if "commercialguru" in host:
                    if not logged_in["commercialguru"]:
                        login_commercialguru(
                            driver=driver,
                            timeout_sec=self.config.timeout_sec,
                            log_callback=self.log,
                            stop_requested=self._should_stop,
                        )
                        logged_in["commercialguru"] = True
                    return self._parse_commercialguru_listing_page

                if not logged_in["propertyguru"]:
                    login_propertyguru(
                        driver=driver,
                        timeout_sec=self.config.timeout_sec,
                        log_callback=self.log,
                        stop_requested=self._should_stop,
                    )
                    logged_in["propertyguru"] = True
                return self._parse_propertyguru_listing_page

            for i, url in enumerate(self.urls, start=1):
                if self._should_stop():
                    self.log("Stop requested. Finishing current listing.")
                    break

                parse_listing = _get_parser_for_url(url)

                self.log(f"[{i}/{total}] Loading: {url}")
                loaded = self._navigate_to_url(driver, url)

                if not loaded:
                    self.error_count += 1
                    row_data: dict = {f: None for f in DETAIL_FIELDS}
                    row_data["URL"] = url
                    self.log(f"[{i}/{total}] FAILED to load: {url}")
                else:
                    row_data = parse_listing(driver.page_source, url)
                    row_data["Agent Phone Number"] = reveal_and_extract_agent_phone(
                        driver=driver,
                        timeout_sec=self.config.timeout_sec,
                        stop_requested=self._should_stop,
                    ) or row_data.get("Agent Phone Number")
                    label = row_data.get("Asking Price") or url
                    self.log(f"[{i}/{total}] OK — {label}")

                ws.append([row_data.get(field) for field in DETAIL_FIELDS])
                processed += 1

                self.progress({
                    "stage": "listings",
                    "current_page": i,
                    "total_pages": total,
                    "listings_processed": processed,
                    "total_listings": total,
                    "error_count": self.error_count,
                    "elapsed": int(time.time() - self.start_time),
                })

            wb.save(self.config.output_csv)
            self.log(f"Saved Excel: {self.config.output_csv}")

        finally:
            try:
                if self._is_driver_alive(driver):
                    driver.quit()
            except Exception:
                pass
            with self._driver_lock:
                self._driver = None

        return {
            "total_pages": total,
            "total_links": total,
            "processed": processed,
            "errors": self.error_count,
            "elapsed": int(time.time() - self.start_time),
        }
